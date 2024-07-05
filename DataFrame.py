import pandas as pd
import numpy as np
import openpyxl
import xlwt
from parameter import Parameter
import gspread #untuk akses google spreadsheet
from gspread_dataframe import set_with_dataframe
import datetime
pm = Parameter()


class dataframe():

    def baca_kct(self, filepath_kct_krn: str, id_pelanggan: int):
        try:
            df = pd.read_excel(filepath_kct_krn)
            # 321610005372
            idpel_search = int(id_pelanggan)
            indexfound = ""
            # get number of rows
            numrows = len(df.index)
            print(numrows)

            for i in range(numrows):
                data = df["IDPEL"][i]
                if (data == idpel_search):
                    indexfound = i
                    break
            kct1a = df["TOKEN_KCT1A"][indexfound]
            kct1b = df["TOKEN_KCT1B"][indexfound]
            kct2a = df["TOKEN_KCT2A"][indexfound]
            kct2b = df["TOKEN_KCT2B"][indexfound]
            petugasbaca = df["PETUGASBACA"][indexfound]
            tanggalbaca = df["TGLBACA"][indexfound]
            statusbaca = df["KETERANGAN_STATUS_BACA"][indexfound]
            message = "Status Baca Idpel "+str(idpel_search) + " : " + statusbaca + \
                "\n" + "Nama Petugas : " + \
                str(petugasbaca)+"\n" + "Tanggal baca : " + str(tanggalbaca) + \
                "\n" + "KCT 1A : " + str(kct1a) + "\n" + "KCT 1B : " + str(
                    kct1b)+"\n" + "KCT 2A : " + str(kct2a) + "\n" + "KCT 2B : " + str(kct2b)
            # message = "Status Baca : " + statusbaca + "\n" + "Nama Petugas : " + petugasbaca, "\n" + "Tanggal baca : " + str(tanggalbaca) + \
            #     "\n" + "KCT 1A : " + str(kct1a) + "\n" + "KCT 1B : " + str(kct1b) + \
            #     "\n" + "KCT 2A : " + str(kct2a) + "\n" + "KCT 2B : " + str(kct2b)
            print("idpelfound : ", indexfound)
            print(message)
            return "yes", message
        except Exception as e:
            message = "Gagal read file excel kct\nMessage Error : "+str(e)
            print(message)
            return "no", message

    def get_userlink_bykodeunit(self, kdunit: str, jenis_user: str, part_link_awal: str, part_link_akhir: str):
        try:
            df = pd.read_excel(pm.filepathlistuser)
            print(df)
            # get nip uSER
            list_username = df[df["KODE_UNIT"] == int(kdunit)]
            username = list_username[list_username["Jabatan"]
                                     == jenis_user]["NIP"]
            password = list_username[list_username["Jabatan"]
                                     == jenis_user]["Password_AP2T"]
            link = part_link_awal + \
                username.item()+part_link_akhir
            # print("Username : ", username, "\nPassword : ", pasword)
            print(link)
            # print("Length user : ", len(username.item()),
            #       "\nLength password : ", len(password.item()), "(", password.item(), ")")
            message = "Berhasil get link dan username - password"
            return link, username.item(), password.item()
        except Exception as e:
            message = "Gagal ambil username, password dan get link, periksa File Excel dan kode unit\Message Error : " + \
                str(e)
            print(message)
            return link, "null", "0"

    def get_authenticated_user(self, filepathlistuser: str):
        try:
            # coba get list user dari database
            df = pd.read_excel(io=filepathlistuser,
                               sheet_name=pm.sheetname_listuserid)
            print(df)
            list_id = df["chat_id"].values.tolist()
            message = "Berhasil get list authenticated user"
            print(message)
            return "yes", list_id, message
        except Exception as e:
            list_id = []
            message = "Lst id gagal di ambil\nMessage Error : "+str(e)
            print(message)
            return "no", list_id, message

    def get_level_user_data(self, chat_id: int):
        try:
            df = pd.read_excel(io=pm.filepathlistuser,
                               sheet_name=pm.sheetname_listuserid)
            selected_row = df[df["chat_id"] == chat_id]
            print("Level user id ", chat_id, " : ",
                  selected_row["level"].item())
            message = "Berhasil get level user"
            return "yes", selected_row["level"].item(), message
        except Exception as e:
            message = "gagal get level user \nMessage Error : "+str(e)
            return "no", "null", message
    # Update Data USer
    def update_user_data(self,filepathlistuser:str,sheetname:str,column_lookup:str,input_value:str,updated_value:str,column_objective:str):
        import openpyxl
        try:
            # Load the existing workbook
            workbook = openpyxl.load_workbook(filepathlistuser)
            worksheet = workbook.get_sheet_by_name(sheetname)
            print("Berhasll")
            #get specified row of input value
            [status,message,row] = dataframe.get_specified_row(workbook=workbook,sheet_name="Sheet1",search_value=input_value)
            if(status == "yes"):
                print("cari kolom update")
                # dataframe.write_to_excel(workbook=workbook,sheet_name=sheetname,row_number=row,column_name=column_objective,value=updated_value)
                # message = "Berhasil update"
                # print(message)
                try:
                    dataframe.write_to_excel(workbook=workbook,sheet_name=sheetname,row_number=row,column_name=column_objective,value=updated_value)
                    message = "Berhasil update"
                    print(message)
                    return "yes", message
                except Exception as e:
                    message = "Gagal update user\nMessage Error : "+str(e)
                    print(message)
                    return "no", message
                
        except Exception as e:
            message = "Gagal Load File listuserid.xls\nMessage Error : \n"+str(e)
            print(message)
    
    def get_specified_row(workbook,sheet_name,search_value):
        # Load the existing workbook
        workbook = workbook

        # Select the desired sheet (e.g., "Sheet1")
        sheet_name = "Sheet1"
        sheet = workbook[sheet_name]

        # Specify the value you want to search for
        search_value = search_value

        # Search for the cell with the given value
        row_position = None
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for cell_value in row:
                if cell_value == search_value:
                    row_position = row_index
                    break

        # Output the row position if found
        if row_position is not None:
            message = f"The value '{search_value}' is found in row {row_position}."
            print(message)
            return "yes",message,row_position
        else:
            message = f"The value '{search_value}' is not found in the sheet."
            print(message)
            return "no",message,None


    def find_column_by_header(workbook, sheet_name, header):
        # Load the Excel workbook
        workbook = workbook

        # Select the desired sheet
        sheet = workbook[sheet_name]

        # Search for the header in the first row of the sheet
        for column in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            for index, cell_value in enumerate(column, start=1):
                if cell_value == header:
                    # If the header is found, return the column letter
                    return openpyxl.utils.get_column_letter(index)
        # If the header is not found, return None
        return None

        if column_letter:
            print(f"The header '{header_to_find}' is located in column {column_letter}.")
        else:
            print(f"The header '{header_to_find}' was not found in the specified sheet.")


    def write_to_excel(workbook, sheet_name, row_number, column_name, value):
        # Load the Excel workbook
        workbook = workbook

        # Select the desired sheet
        sheet = workbook[sheet_name]

        # Write the value to the specified cell
        sheet[column_name + str(row_number)] = value

        # Save the changes to the Excel file
        workbook.save(pm.filepathlistuser)  

    def get_kode_unit_user(self, chat_id: int):
        try:
            df = pd.read_excel(io=pm.filepathlistuser,
                            sheet_name=pm.sheetname_listuserid)
            selected_row = df[df["chat_id"] == chat_id]
            print(selected_row)
            print("kode unit selected : ", selected_row["kode_unit"].item())
            message = "Berhasil get kode unit chat id"
            print(message)
            return "yes", selected_row["kode_unit"].item(), message
        except Exception as e:
            message = "Gagal mengambil kode unit user\nMessage Error : "+str(e)
            return "no", "null", message
    def get_kode_unit_user_tagsus(self, chat_id: int):
        try:
            df = pd.read_excel(io=pm.filepathlistuser,
                               sheet_name=pm.sheetname_listuserid)
            selected_row = df[df["chat_id"] == chat_id]
            print(selected_row)
            print("kode unit selected : ", selected_row["up3"].item())
            message = "Berhasil get kode unit chat id"
            print(message)
            return "yes", selected_row["up3"].item(), message
        except Exception as e:
            message = "Gagal mengambil kode unit user\nMessage Error : "+str(e)
            return "no", "null", message

    def get_convert_kdunit_to_nama_unit(self,kdunit_selected):
        try:
            df = pd.read_excel(io="data\\listuser\\listkodeunit.xlsx", sheet_name="kdunit")
            selected_row = df[df["kodeunit"] == kdunit_selected]
            print("Nama unit selected : ", selected_row["namaunit"].item())
            print(selected_row)
            message = "Berhasil get nama unit"
            return "yes",selected_row["namaunit"].item(),message
        except Exception as e:
            message = "Gagal menemukan nama unit dari kode unit\nError Message : "
            print(message,str(e))
            return "no","null",message+str(e)
    def get_monthname_from_number(self,number_code:str):
        months = np.array(['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
                   'Agustus', 'September', 'Oktober', 'November', 'Desember'])
        selected_month = months[int(number_code) - 1]
        return selected_month

    def get_user_acmt(self):
        try:
            df = pd.read_excel(io=pm.filepathlistuser, sheet_name="Sheet1")
            selected_row = df[df["Akun"] == "ACMT"]
            username = selected_row["NIP"].item()
            password = selected_row["Password_AP2T"].item()
            print("Username : ", username, "Password : ", password)
            return "yes", username, password
        except Exception as e:
            message = "Gagal read list user acmt\nMessage Error : "+str(e)
            print(message)
            return "no", "null", "null"

    def get_jumlah_klik_resetimei(self, kodeunit_cari: str):
        # Jumlah klik tombol down di sesuaikan dengan nomor letak ULP nya pada dropdown AP2T reset imei
        df = pd.read_excel(io=pm.filepathlistuser, sheet_name="listulp")
        rows_klik = df[df["kode_unit"] == kodeunit_cari]
        jumlah_perulangan = rows_klik["no"].item()
        print("jumlah klik : ", jumlah_perulangan)
        print(type(jumlah_perulangan))
        return int(jumlah_perulangan)

    def get_last_row(self, filepathlistuser: str, sheetname: str):
        try:
            df = pd.read_excel(io=filepathlistuser,
                               sheet_name=sheetname)
            print("Baris akhir pada excel : ", str(len(df["no"])+1))
            nomor_rows = len(df["no"])+2
            message = "Berhasil get last row"
            print(message)
            return "yes", nomor_rows, message
        except Exception as e:
            message = "Gagal get last rows, harap periksa file sedang tidak digunakan" + \
                str(e)
            print(message)
            return "no", 0, message

    def write_data_userid(self, filepathlistuser: str, sheetname: str, row: int, chat_id: int = 0, kode_unit: int = 0, nama: str = "", nomor_telfon: str = "", level: str = "user"):
        # select the workbook
        try:
            workbook = openpyxl.load_workbook(filename=filepathlistuser)
            worksheet = workbook.get_sheet_by_name(sheetname)
            try:
                worksheet.cell(row=row, column=1).value = row-1
                worksheet.cell(row=row, column=2).value = chat_id
                worksheet.cell(row=row, column=3).value = kode_unit
                worksheet.cell(row=row, column=4).value = nama
                worksheet.cell(row=row, column=5).value = nomor_telfon
                worksheet.cell(row=row, column=6).value = level
                workbook.save(filepathlistuser)
                message = "Berhasil add user"
                print(message)
                return "yes", message
            except Exception as e:
                message = "Gagal add user\nMessage Error : "+str(e)
                print(message)
                return "no", message
        except Exception as e:
            message = "Gagal load excel\n"+str(e)
            print(message)
            return "no", message

    def read_reportservlet(skip_rows :int = 8):
        pm = Parameter()
        namafiletagsus =  pm.download_ts+"\\"+pm.namafiletagsus
        try:
            df = pd.read_excel(io="data\\downloads\\ReportServlet.xls",sheet_name="Report1",skiprows=skip_rows)
            message = "Berhasil baca ReportServlet"
            print(message)
            return "yes",message,df
        except Exception as e:
            message = "Gagal baca ReportServlet\nMessage Error : \n"+str(e)
            print(message)
            return "no", message, "null"
    def read_laporan_pd():
        namafile =  'data\\downloads\\EIS\\GV.xls'
        try:
            df = pd.read_excel(io=namafile,sheet_name="Sheet",skiprows=0)
            message = "Berhasil baca ReportServlet"
            print(message)
            return "yes",message,df
        except Exception as e:
            message = "Gagal baca ReportServlet\nMessage Error : \n"+str(e)
            print(message)
            return "no", message, "null"
# print(df.loc[:,'2'])

    def log_data(self, chat_id: int, activity: str, time: str):
        pm = Parameter()
        try:
            status, last_row, message = self.get_last_row(
                filepathlistuser=pm.filepathlog, sheetname="log")
            if (status == "yes"):
                print("Last row : "+str(last_row))
                print(message)
                # Write number of rows
                workbook = openpyxl.load_workbook(filename=pm.filepathlog)
                worksheet = workbook.get_sheet_by_name("log")
                worksheet.cell(row=last_row, column=1).value = last_row-1
                worksheet.cell(row=last_row, column=2).value = chat_id
                worksheet.cell(row=last_row, column=3).value = activity
                worksheet.cell(row=last_row, column=4).value = time
                workbook.save(pm.filepathlog)
            else:
                print("Gagal write log data")
        except Exception as e:
            print("Gagal write log\nMesage Error : ", str(e))

    @staticmethod
    def get_all_userid(filepathlistuserid: str, level_user: str):
        df = pd.read_excel(io=filepathlistuserid, sheet_name="listuserid")
        # print(df)
        return df["chat_id"]

#Write Data to Google Sheet
    def write_df_to_google_sheet(filepathjson,GSHEET,TAB_NAME,df,first_row,first_col):
        #access the spreadsheet
        try:
            gc = gspread.service_account(filepathjson)
            sh = gc.open(GSHEET)
            worksheet = sh.worksheet(TAB_NAME)
            set_with_dataframe(worksheet=worksheet,dataframe=df,row=first_row,col=first_col)
            message = "Berhasil write data ke Google Spreadsheet"
            return "yes",message
        except Exception as e:
            message = "Gagal Write data ke Google Spreadsheet\nMessage Error : \n"+str(e)
            return "no",message
    def read_from_googlesheet_to_df(filepathjson,GSHEET,TAB_NAME,Cell:str):
        gc = gspread.service_account(filename=filepathjson)
        sh = gc.open(GSHEET)
        worksheet = sh.worksheet(TAB_NAME)
        val = worksheet.acell(Cell).value
        return val
    def get_nama_sheet_bulan_sekarang():
        # Get the current date and time
        current_date = datetime.datetime.now()
        print(str(current_date.day))
        print(str(current_date.year))
        # Extract the month from the current date
        current_month = current_date.month
        print(str(current_month))
        #array of sheet names
        months = np.array(['1.Januari', '2.Februari', '3.Maret', '4.April', '5.Mei', '6.Juni', '7.Juli',
                   '8.Agustus', '9.September', '10.Oktober', '11.November', '12.Desember'])
        selected_month = months[current_month - 1]
        print(selected_month)
        #return yearnow,monthnow,datenow
        return selected_month
    
    def get_tahun_bulan_sekarang():
        # Get the current date and time
        current_date = datetime.datetime.now()
        # Extract the month from the current date
        current_month = 0
        if(current_date.month<10):
            current_month = "0"+str(current_date.month)
        else:
            current_month = str(current_date.month)

        tahun_bulan = str(current_date.year)+current_month
        print(tahun_bulan)
        #return yearnow,monthnow,datenow
        return tahun_bulan

